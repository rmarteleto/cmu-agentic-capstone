"""SIG spreadsheet I/O — reads historical SIG workbooks for corpus ingestion."""
from __future__ import annotations


def read_qa_pairs(path: str, question_column: str = "Question",
                  answer_column: str = "Answer") -> list[dict]:
    """Read a historical (filled) SIG workbook into {'question','answer'} rows.

    Feeds the SIG-history corpus (candidate answers). Column names are matched
    case-insensitively; falls back to the first two columns when headers are
    absent. Rows missing either side are skipped.
    """
    import openpyxl
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    header = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]

    def _col(name: str, default: int) -> int:
        try:
            return header.index(name.lower())
        except ValueError:
            return default

    qcol, acol = _col(question_column, 0), _col(answer_column, 1)
    ccol = _col("comment", -1)
    rows: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        q = row[qcol] if qcol < len(row) else None
        a = row[acol] if acol < len(row) else None
        c = row[ccol] if ccol >= 0 and ccol < len(row) else None
        if q and a:
            rows.append({
                "question": str(q).strip(),
                "answer": str(a).strip(),
                "comment": str(c).strip() if c else "",
            })
    return rows
